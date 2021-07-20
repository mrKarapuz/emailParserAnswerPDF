
from aiogram import Bot
import imaplib, smtplib, os, pdfplumber, MySQLdb, requests
from time import sleep, strftime
from openpyxl import load_workbook
from zipfile import ZipFile
from email import message_from_bytes
from email.header import decode_header, make_header
from pathlib import Path
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

print('СЕРВЕР ЗАПУЩЕН')

SYMVOL = ('<', '>', ':', '"', '/', '\\', '|', '?', '*')

# Функция отправки сообщения в группу в телеграме
def telegram_bot_send():
    token = '1901007184:AAHZ4DCUsqD9MRCzAdLYCgrHscz6erCy4cU'
    bot = Bot(token=token)
    chat_id = '-1001572207430'
    text = win_code
    url = f'https://api.telegram.org/bot{token}/sendMessage?chat_id={chat_id}&parse_mode=Markdown&text={text}'
    requests.get(url)

# Функция, которая срабатывает если не подключились к почте или не подключились к базе данных
def error_and_exit():
    print('---------------------')
    print(f'Выход через 25 сек...')
    sleep(10)
    print(f'Выход через 15 сек...')
    sleep(10)
    print(f'Выход через 5 сек...')
    sleep(5)
    exit()

# Функция загрузки файлов
def get_attachments(msg):
    global myzip
    try:
        with ZipFile(zip_dir +'.zip', 'w') as myzip:
            for part in msg.walk():
                if part.get_content_maintype()=='multipart':
                    continue
                if part.get('Content-Disposition') is None:
                    continue
                fileName = part.get_filename()
                if bool(fileName):
                    filePath = os.path.join(str(make_header(decode_header(fileName))))
                    with open(filePath,'wb') as f:
                        f.write(part.get_payload(decode=True))
                        myzip.write(filePath)
                    os.remove(filePath)
    except FileNotFoundError:
            print(f'''Не удается найти папку для загрузки отработанных файлов.
Проверьте данные и перезапустите программу.''')
            error_and_exit()

# Функция поиска Win кода на сервере
def serach_win_code_in_file(_win_code):
    global filename
    try:
        for elem in os.listdir(DIR_OF_SERVER):
            if len(elem) > 20:
                filename = DIR_OF_SERVER + elem
                text = ''
                pages = []
                with pdfplumber.open(filename) as pdf:
                    for i, page in enumerate(pdf.pages):
                        text = text+str(page.extract_text())
                        pages.append(page)

                win_code_in_file = text[(text.find('номер кузова:')) + 14 :(text.find('номер кузова:')) + 31]
                if _win_code == win_code_in_file : return True
            else: continue
    except FileNotFoundError:
        print(f'''Не удается найти путь к серверу: 
{DIR_OF_SERVER}  
Проверьте данные и перезапустите программу.''')
        error_and_exit()

# Функция отправки сообщения пользователю
def send_message_to_user(email_user):
    message = MIMEMultipart()
    message['From'] = LOGIN
    message['To'] = email_user
    message['Subject'] = 'Re:' + subject_of_mail
    if TEXT_MESSAGE:
        body = TEXT_MESSAGE
    else:
        body = '''Отримайте Вашу довідку'''

    name_to_file_on_mail = filename[filename.rfind('\\') + 1 :]
    message.attach(MIMEText(body, 'plain'))
    pdfname = filename
    binary_pdf = open(pdfname, 'rb')
    payload = MIMEBase('application', 'octate-stream', Name=name_to_file_on_mail)
    payload.set_payload((binary_pdf).read())
    encoders.encode_base64(payload)
    payload.add_header('Content-Decomposition', 'attachment', filename=pdfname)
    message.attach(payload)
    TYPE_OF_EMAIL = 'smtp.' + LOGIN[LOGIN.find('@')+1:]
    session = smtplib.SMTP(TYPE_OF_EMAIL, 587)
    session.starttls()
    session.login(LOGIN, PASSWORD)
    text = message.as_string()
    session.sendmail(LOGIN, email_user, text)
    session.quit()
# Считывание данных с базы данных
try:
    file_xlsx = load_workbook(filename = 'data.xlsx')
    sheet = file_xlsx['settings']
    # Логин электронной почты
    LOGIN = sheet.cell(row=1, column=2).value
    # Пароль электронной почты
    PASSWORD = sheet.cell(row=2, column=2).value
    # Имя папки, из которой будут считываться письма, по умолчанию "Входящие"
    NAME_OF_CATALOG_ON_MAIL = sheet.cell(row=3, column=2).value
    # Папка сервера
    DIR_OF_SERVER = str(sheet.cell(row=4, column=2).value + '\\')
    # Папка отработанных данных
    DIR_OF_SERVER_DONE = str(sheet.cell(row=5, column=2).value + '\\')
    # Путь к папке сохранения архивов
    ATTACHMENT_DIR = str(sheet.cell(row=6, column=2).value + '\\')
    # Время перезагрузки сервера
    TIME_TO_RELOAD = sheet.cell(row=7, column=2).value
    # Текст сообщения, которое будет отправлено пользователю
    TEXT_MESSAGE = sheet.cell(row=8, column=2).value
    
    file_xlsx.close()
except FileNotFoundError: 
    print(f'''Не удается найти файл базы данных, убедитесь что файл называется "data.xlsx" и находится в одной папке с файлом программы
Проверьте данные и перезапустите программу.''')
    error_and_exit()

# Подключение к почте
print(f'Подключение к почтовому ящику {LOGIN} ...')
try:
    TYPE_OF_EMAIL = 'imap.' + LOGIN[LOGIN.find('@')+1:]
    con = imaplib.IMAP4_SSL(TYPE_OF_EMAIL)
    con.login(LOGIN, PASSWORD)
except:
    print(f'''Не удается подключится к почтовому ящику, проверьте плавильность введенных вами данных в файле "data.xlsx", лист "settings":  
    Логин: {LOGIN}
    Пароль: {PASSWORD}
Также проверьте что в настройках вашей почты включена функция доступа с помощью IMAP и разрешен доступ "небезопасных приложений" (для почты Gmail.com)''')
    error_and_exit()
# Проверка правильности введеного времени перезагрузки сервера
if not type(TIME_TO_RELOAD) == int:
    print()
    print('Не корректно указано время, через которое сервер будет автоматически проверять почту. Сервер будет обновляться каждые 30 сек.')
    TIME_TO_RELOAD = 30
print(f'''
Успешно подключено к почтовому ящику {LOGIN}
Папка загрузки архивов: {ATTACHMENT_DIR}
Путь к серверу: {DIR_OF_SERVER}
Путь к отработанным данным: {DIR_OF_SERVER_DONE}
Интервал автоматической перезагрузки: {TIME_TO_RELOAD} сек.
Время: {strftime('%H:%M:%S')}
''')


# Счетчик не занятых ячеек для записи в базе данных
count_of_xlsx = 2

# Автоматическое сканирование почты
while True:
    # Количество не обработанных сообщений на почте
    count_of_messages = int(con.select(NAME_OF_CATALOG_ON_MAIL)[1][0].decode('utf-8'))
    # Загружаем каждое сообщение по очереди
    for i in range(1, count_of_messages + 1):
        result, data = con.fetch(f'{i}', '(RFC822)')
        raw = message_from_bytes(data[0][1])
        # Имя и адрес отправителя
        from_user = str(make_header(decode_header(raw['FROM'])))
        # Тема письма
        subject_of_mail = str(make_header(decode_header(raw['Subject'])))
        # Проверяем тему письма на запрещенные для сохранения символы
        for elem in SYMVOL:
            if elem in subject_of_mail:
                subject_of_mail = subject_of_mail.replace(elem, '')
        # Объявляем название архива
        
            zip_dir = os.path.join(ATTACHMENT_DIR, str(subject_of_mail))
            
        # Находим WIN-code машины в теме письма
        win_code = False
        for elem in subject_of_mail.split(' '):
            if len(elem) == 17 and (not 'i' 'o' 'q' in elem.lower()) and elem.isalnum():
                win_code = elem.upper()
        # Если win code находится в одном из файлов
        if serach_win_code_in_file(win_code):
            # Сохраняем архив
            get_attachments(raw)
            # Если архив не пуст
            if len(myzip.infolist()) != 0:
                # Убираем в конце названия отправляемого файла имя и фамилию
                os.rename(filename, filename[:filename.rfind('_')] + '.pdf')
                filename = filename[:filename.rfind('_')] + '.pdf'
                # Отправляем сообщение пользователю
                send_message_to_user(from_user[from_user.find('<') + 1 :from_user.find('>')])
                print(f'Сообщение успешно обработано от пользователя: {from_user}', strftime('%b, %A, %H:%M:%S'))
                # Перемещаем файл в папку server_done
                try:
                    os.replace(filename, DIR_OF_SERVER_DONE + filename[filename.rfind('\\'):])
                except FileNotFoundError:
                    print(f'''Не удается найти путь к папке: 
{DIR_OF_SERVER_DONE}  
Файл не был перемещен
Проверьте данные и перезапустите программу.''')
                    continue
                # Отправляем сообщение в группу телеграм
                telegram_bot_send()
                # Присваиваем сообщению флаг для его дальнейшего удаления (IMAP пометит как удаленное, на почте отправится в архив)
                con.store(f'{i}', '+FLAGS', '\\Deleted')
                # Работа с базой данных
                db = MySQLdb.connect(host="176.111.49.48",    
                     user="zkdqsgeo_euro",        
                     passwd="M8s4J5j2",     
                     db="zkdqsgeo_euro")
                cur = db.cursor()
                Date = strftime('%Y-%m-%d')
                Email = from_user[from_user.find('<') + 1 :from_user.find('>')]
                cur.execute(f'INSERT INTO basa (date, email, win_code) VALUES(%s, %s, %s)', (Date, Email, win_code))
                db.commit()
                db.close()
            else:
                os.remove(zip_dir + '.zip')
    con.expunge()        
    count_of_messages = int(con.select(NAME_OF_CATALOG_ON_MAIL)[1][0].decode('utf-8'))        
    print(f'Не обработанных сообщений на почте: {count_of_messages}')
    sleep(TIME_TO_RELOAD)